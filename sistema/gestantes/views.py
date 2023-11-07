
from django.utils import timezone 
import os
from django.shortcuts import get_object_or_404, render, redirect
from .forms import DatosGestanteForm,ActividadGestanteForm,SeguimientoGestanteForm,SoporteGestanteForm, CargueMensualForm
from .models import DatosGestante, ActividadGestante, SoporteGestante, SeguimientoGestante,InformacionGestante,CargueMensual, InformacionGestante, InformacionMensual
from django.http import FileResponse, HttpResponseBadRequest, HttpResponseRedirect, HttpResponse, JsonResponse
import requests
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
def home(request):
    return render(request, "home.html") 

def semanal(request):
    return render(request, "plantillas/prestador_semanal") 


def registros(request):
    num_distintos = DatosGestante.objects.values('nro_doc','tipdoc').distinct()
    #print(num_distintos)
    context = {
 
        'num_distintos': num_distintos,
    }
    return render(request, "plantillas/registros_general.html", context)


def asociados(request, tipo_cedula, numero_cedula):
    registros = DatosGestante.objects.filter(tipdoc=tipo_cedula, nro_doc=numero_cedula)
    registros_soporte = SoporteGestante.objects.filter(datosgestante__in=registros).distinct()
    context = {
        'registros_completos': registros_soporte,  
        'tipo_cedula': tipo_cedula,
        'numero_cedula': numero_cedula,
    }

    return render(request, 'plantillas/registros_individual.html', context)



def generar_ruta_soporte(datos_gestante, soporte,regional,prestador):
    ruta = 'SOPORTES_GESTANTES'
    filename = soporte.soporte.name#campo e la tabla
    instance = soporte
    instance.regional = regional
    instance.prestador = prestador
    instance.tipdoc = datos_gestante.tipdoc
    instance.identificacion = datos_gestante.nro_doc
    instance.tipoar = soporte.tipo_arch
    soporte.soporte.name = os.path.join(ruta, instance.regional, instance.prestador, instance.tipdoc, instance.identificacion, instance.tipoar, filename)



@login_required
def formulario(request):
    if request.method == 'POST':
        form_datos_gestante = DatosGestanteForm(request.POST)
        form_actividades = ActividadGestanteForm(request.POST)
        form_seguimiento = SeguimientoGestanteForm(request.POST)
        form_soportes = SoporteGestanteForm(request.POST, request.FILES)

        if form_datos_gestante.is_valid():
            datos_gestante = form_datos_gestante.save(commit=False)
            if isinstance(request.user, User):  # Verificar que request.user sea una instancia válida de User
                datos_gestante.usuario = request.user
                datos_gestante.save()  #hacer else si no esta registrado

            if form_actividades.is_valid():
                actividades = form_actividades.save(commit=False)
                actividades.datosgestante = datos_gestante
                actividades.save()

            if form_seguimiento.is_valid():
                seguimiento = form_seguimiento.save(commit=False)
                seguimiento.datosgestante = datos_gestante
                seguimiento.save()

            if form_soportes.is_valid():
                regional = request.POST.get('regional')
                prestador = request.POST.get('prestador')
                soportes = form_soportes.save(commit=False)
                soportes.datosgestante = datos_gestante
                generar_ruta_soporte(datos_gestante, soportes, regional, prestador)
                soportes.save()
            #limpiar campos
            form_datos_gestante = DatosGestanteForm()
            form_actividades = ActividadGestanteForm()
            form_seguimiento = SeguimientoGestanteForm()
            form_soportes = SoporteGestanteForm()
            return redirect('formulario') #urls
                
   
    else:
        form_datos_gestante = DatosGestanteForm()
        form_actividades = ActividadGestanteForm()
        form_seguimiento = SeguimientoGestanteForm()
        form_soportes = SoporteGestanteForm()
        
    context = {
        'form_datos_gestante': form_datos_gestante,
        'form_actividades': form_actividades,
        'form_seguimiento': form_seguimiento,
        'form_soportes': form_soportes,
    }
    return render(request, 'plantillas/formulario.html', context)




def datos(request):
    datos_gestantes = DatosGestante.objects.prefetch_related('actividadgestante_set', 'seguimientogestante_set', 'soportegestante_set').order_by('-created').all()

    data_list = []
    for gestante in datos_gestantes:
        actividades_list = gestante.actividadgestante_set.all()
        seguimientos_list = gestante.seguimientogestante_set.all()
        soportes_list = gestante.soportegestante_set.all()

        actividades_data = actividades_list[0] if actividades_list else None
        seguimientos_data = seguimientos_list[0] if seguimientos_list else None
        soportes_data = soportes_list[0] if soportes_list else None
        """ precisar la consulta """
        data_dict = {
            'id': gestante.id,
            'tipdoc': gestante.tipdoc,
            'nro_doc': gestante.nro_doc,
            'created': gestante.created,
            'actividades': actividades_data,
            'seguimientos': seguimientos_data,
            'soportes': soportes_list,
        }
        data_list.append(data_dict)
        """ print(data_list) """


    context = {
        'data_list': data_list,
    }
    return render(request, 'plantillas/semanal.html', context)





from django.db import models



from django.db import IntegrityError
import pandas as pd

def procesar_excel(file_path, mensual):
    # Crear un diccionario para mapear verbose_name a nombre de campo
    field_name_mapping = {field.verbose_name: field.name for field in InformacionMensual._meta.fields}

    excel_file = pd.ExcelFile(file_path)
    sheet_name = excel_file.sheet_names[2]
    df = pd.read_excel(file_path, sheet_name)
    for index, row in df.iterrows():
        if (
        index in [0, 1] 
            ):
            continue
       
        row_dict = row.to_dict()
        updated_row_dict = {}

        for key, value in row_dict.items():
            field_name = field_name_mapping.get(key)

            if field_name is not None:
                updated_row_dict[field_name] = value
            else:
                print(f'Campo no válido en la fila {index + 1}, columna "{key}": {value}')
                continue

        try:
            informacion_gestante = InformacionMensual(**updated_row_dict)
            informacion_gestante.mensual = mensual
            informacion_gestante.save()
        except Exception as e:
            print(f'Error al guardar en la fila {index + 1}: {e}')




#guardar y procesar excel mensual
def mensual(request):
    #listar
    archivos_excel = CargueMensual.objects.order_by('-created').all()
    #guardar
    if request.method == 'POST':
        form = CargueMensualForm(request.POST, request.FILES)
        if form.is_valid():
            form.instance.soporte_excel.name = generar_ruta_cargue_mensual(form.instance, request.FILES['soporte_excel'].name)
            form.save()
            procesar_excel(form.instance.soporte_excel.path, form.instance)
            return redirect('mensual') 
    else:
        form = CargueMensualForm()
    context = {
        'form': form,
        'archivos_excel': archivos_excel,
    }
    return render(request, "plantillas/mensual.html", context)



def generar_ruta_cargue_mensual(instance, filename):
    ruta = 'SOPORTES_MENSUAL_GESTANTES'
    año = timezone.now().year 
    prestador = "INTEGRAL"
    mes = timezone.now().strftime('%B')
    return os.path.join(ruta, str(año), prestador, mes, filename)








from openpyxl import Workbook
from django.apps import apps
from datetime import datetime

def generar_excel_mensual(request):
    informacion_gestantes = InformacionMensual.objects.all()

    wb = Workbook()
    ws = wb.active

    excluded_fields = ['mensual']

    field_names = [field.name for field in InformacionMensual._meta.fields if field.name not in excluded_fields]
    ws.append(field_names)

    for informacion_gestante in informacion_gestantes:
        row_data = [str(getattr(informacion_gestante, field)) for field in field_names]
        ws.append(row_data)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="informe_mensual.xlsx"'
    wb.save(response)

    return response






import openpyxl
from django.conf import settings

def generar_excel_semanal(request):
    file_path = os.path.join(settings.BASE_DIR, 'gestantes', 'static', 'excel', 'baseg.xlsx')
    print(file_path)

    print(file_path)
    wb = openpyxl.load_workbook(file_path)

    
    ws_datos_gestante = wb['REGISTRO 2']
    ws_actividad_gestante = wb['RESGISTRO 3']
    ws_seguimiento_gestante = wb['REGISTRO 4']

    datos_gestante = DatosGestante.objects.all()
    actividad_gestante = ActividadGestante.objects.all()
    seguimiento_gestante = SeguimientoGestante.objects.all()


    row_num = 5
    for dg in datos_gestante:
        ws_datos_gestante.cell(row=row_num, column=1, value=dg.tipdoc)
        row_num += 1

    row_num = 5
    for ag in actividad_gestante:
        ws_actividad_gestante.cell(row=row_num, column=1, value=dg.tipdoc)
        ws_actividad_gestante.cell(row=row_num, column=2, value=dg.nro_doc)
        ws_actividad_gestante.cell(row=row_num, column=3, value=dg.nro_doc)
        row_num += 1


    row_num = 5
    for sg in seguimiento_gestante:
        ws_seguimiento_gestante.cell(row=row_num, column=1, value=dg.tipdoc)
        ws_seguimiento_gestante.cell(row=row_num, column=2, value=dg.nro_doc)
        ws_seguimiento_gestante.cell(row=row_num, column=3, value=sg.tipo_caso) 
        ws_seguimiento_gestante.cell(row=row_num, column=12, value=sg.info_ciudado_sal)
        row_num += 1

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="InformeGeneral.xlsx"'
    wb.save(response)

    return response


def eliminar_mensual(request, datosgestante_id):
    datos_gestante = get_object_or_404(CargueMensual, pk=datosgestante_id)
    datos_gestante.informacion_gestantes.all().delete()
    datos_gestante.delete()
    return redirect('mensual')
